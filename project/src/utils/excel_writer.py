"""
Utility to fill the test_template.xlsx with predictions.
"""

from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl


def fill_test_template(
    predictions_future: np.ndarray,
    predictions_missing: dict,
    template_path: Optional[Path] = None,
    output_path: Optional[Path] = None,
) -> Path:
    """
    Fill the test template Excel file with predictions.

    Parameters
    ----------
    predictions_future : (6, 224) array
        Predicted prices for the 6 future rows.
    predictions_missing : dict[int, np.ndarray]
        Maps row index (0-based in template data) to a dict:
            {col_index: predicted_value} for the NA columns.
        Or alternatively: {row_idx: (224,) full predicted vector}.
    template_path : path to test_template.xlsx
    output_path : where to save the filled file

    Returns
    -------
    output_path : Path
    """
    data_dir = Path(__file__).resolve().parents[2] / "data"
    if template_path is None:
        template_path = data_dir / "test_template.xlsx"
    if output_path is None:
        output_path = data_dir / "results.xlsx"

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Discover template row types dynamically from the first column ("Type").
    future_rows = []
    for excel_row in range(2, ws.max_row + 1):
        row_type = str(ws.cell(row=excel_row, column=1).value or "").lower()
        if "future" in row_type:
            future_rows.append(excel_row)

    if len(future_rows) != len(predictions_future):
        raise ValueError(
            f"predictions_future has {len(predictions_future)} rows, "
            f"but template has {len(future_rows)} future rows."
        )

    for i, excel_row in enumerate(future_rows):
        surface = predictions_future[i]
        for j in range(224):
            excel_col = j + 2  # col 2 is first price col (col 1 is Type)
            ws.cell(row=excel_row, column=excel_col, value=float(surface[j]))

    # Missing data rows are indexed in predictions_missing by template data index (0-based).
    for data_idx, predictions in predictions_missing.items():
        excel_row = data_idx + 2
        if isinstance(predictions, np.ndarray) and predictions.shape == (224,):
            # Full vector — only fill the NA cells
            for j in range(224):
                cell = ws.cell(row=excel_row, column=j + 2)
                if cell.value == "NA" or cell.value is None:
                    cell.value = float(predictions[j])
        elif isinstance(predictions, dict):
            # Dict of {col_idx: value}
            for col_idx, val in predictions.items():
                ws.cell(row=excel_row, column=col_idx + 2, value=float(val))

    wb.save(output_path)
    return output_path
